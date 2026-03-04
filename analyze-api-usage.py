import openpyxl
import json

# Load the spreadsheet
wb = openpyxl.load_workbook('/home/ubuntu/rental-calculator/api-usage-analysis.xlsx')
ws = wb.active

# Print all sheet names
print("Sheet names:", wb.sheetnames)
print()

# Print headers
headers = []
for cell in ws[1]:
    headers.append(cell.value)
print("Headers:", headers)
print()

# Print all rows
print(f"Total rows (excluding header): {ws.max_row - 1}")
print(f"Total columns: {ws.max_column}")
print()

# Print all data
for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
    if i == 0:
        print("HEADER:", row)
    else:
        print(f"Row {i}:", row)
    if i > 100:
        print("... (truncated)")
        break
